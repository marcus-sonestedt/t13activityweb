import logging
import calendar
import locale
import datetime
import openpyxl

from django.core.management.base import BaseCommand, CommandError
from app.models import Activity, ActivityType, Event, EventType, Member


def exportScheduleToExcel(filename, year):
    wb = openpyxl.Workbook(iso_dates=True)
    ws = wb._sheets[0]

    driver_type = ActivityType.objects.get(name__startswith="Förare")
    cols = {}

    c = 1
    for t in ["Aktivitet", "Datum", "Vecka", "Dag",
              "Uppgift", "Start", "Slut",
              "Ersättning", "Mat", "Hyrkart",
              "Publ. datum", "Koordinator",
              "Guldkortsnr", "Namn", "Mobilnummer", "Kommentar"]:
        cols[t] = c
        ws.cell(1, c).value = t
        ws.column_dimensions[chr(64+c)].best_fit = True
        c += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['E'].width = 25
    for c in ['B','C','D','K']:
        ws.column_dimensions[c].width = 10
    for c in ['F','G']:
        ws.column_dimensions[c].width = 7
    for c in ['H','I','J']:
        ws.column_dimensions[c].width = 4

    header_font = openpyxl.styles.Font(bold=True)
    computed_font = openpyxl.styles.Font(italic=True)

    for cell in ws["1:1"]:
        cell.font = header_font

    r = 3
    for event in Event.objects.filter(start_date__year=year):
        print(f'{event.start_date}: {event.name}')

        ws.cell(r, 1).value = event.name

        datecell = ws.cell(r, 2)
        datecell.value = event.start_date
        datecell.number_format = 'YYYY-MM-DD'

        weekcell = ws.cell(r, 3)
        weekcell.value = f'=ISOWEEKNUM({datecell.coordinate})' 
        weekcell.data_type = 'f'

        daycell = ws.cell(r, 4)
        daycell.value = f'={datecell.coordinate}' 
        daycell.number_format = 'ddd'

        for c in [weekcell, daycell]:
            c.font = computed_font

        ws.cell(r, cols['Koordinator']).value = ", ".join(
            c.get_fullname() for c in event.coordinators.all())

        for a in event.activities.exclude(type__in=[driver_type]).order_by('type__id'):
            col = [a.start_time, a.end_time]
            for i, t in enumerate(col):
                ws.cell(r, cols['Start'] + i).value = t
                ws.cell(r, cols['Start'] + i).number_format = 'hh:mm'

            ws.cell(r, cols['Publ. datum']).value = a.earliest_bookable_date

            if a.type:
                ws.cell(r, cols['Uppgift']).value = a.type.name
                for i, t in enumerate([ a.type.fee_reimbursed,
                                   a.type.food_included, a.type.rental_kart]):
                    ws.cell(r, cols['Ersättning'] + i).value = 'Ja' if t else 'Nej' 

            r += 1


    with open(filename, 'wb') as file:
        wb.save(file)

    print(f"Exported {r-2} activities to: {filename}")

class Command(BaseCommand):
    help = "Exports a year's schedule to an MS Excel file"

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help="Excel file path")
        parser.add_argument('--year', type=int, help="Year of data")

    def handle(self, *args, **options):
        year = options['year'] or datetime.date.today().year
        outfile = options['file'] or f'Aktivitetslista T13 {year}.xlsx'

        exportScheduleToExcel(outfile, year)
