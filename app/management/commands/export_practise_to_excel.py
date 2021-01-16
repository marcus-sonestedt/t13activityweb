import logging
import calendar
import locale
import datetime
import openpyxl

from django.core.management.base import BaseCommand, CommandError
from app.models import Activity, ActivityType, Event, EventType, Member

logger = logging.getLogger(__name__)

def exportPractiseSessions(file, year):
    wb = openpyxl.Workbook(iso_dates=True)
    sheet = wb._sheets[0]

    flaggvakt_type = ActivityType.objects.get(name="Träningsvakt (klubbträning)")
    pitguide_type = ActivityType.objects.get(name="Depå-fadder")

    sheet.cell(1, 1).value = "Datum"
    sheet.cell(1, 2).value = "Tid"
    sheet.cell(1, 3).value = flaggvakt_type.name
    sheet.cell(1, 4).value = pitguide_type.name

    def activity_assigned(event, start_time, type):
        activities = event.activities.filter(start_time=start_time, type=type)
        if not activities:
            return ''

        return ', '.join((m.assigned.get_fullname() if m.assigned else '') for m in activities)

    row = 2
    for event in Event.objects.filter(start_date__year=year, type__name="Träning"):
        for a in event.activities.filter(type__in=[flaggvakt_type, pitguide_type]).order_by('start_time'):
            t = f'{a.start_time} - {a.end_time}'
            if sheet.cell(row, 2).value != t:
                row += 1
                sheet.cell(row, 1).value = event.date

            sheet.cell(row, 2).value = t
            sheet.cell(row, 3).value = activity_assigned(event, a.start_time, flaggvakt_type)
            sheet.cell(row, 4).value = activity_assigned(event, a.start_time, pitguide_type)

    for cd in sheet.column_dimensions:
        cd.bestFit = True

    wb.save(file)

    print(f"Exported {row - 2} practise event(s) for year {year} to {file}.")

class Command(BaseCommand):
    help = 'Export coordinator and support for all practise sessions for a year'

    def add_arguments(self, parser):
        parser.add_argument('file', type=str, help="Excel file path")
        parser.add_argument('--year', type=int, help="Year of data")

    def handle(self, *args, **options):
        with open(options['file'], 'wb') as file:
            exportPractiseSessions(file, options['year'])