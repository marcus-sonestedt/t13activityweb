import logging
import calendar
import locale
import datetime
import openpyxl

from django.core.exceptions import PermissionDenied
from app.models import Event, EventType, Activity, ActivityType, Member

logger = logging.getLogger(__name__)

class ExcelSheet:
    '''Defines structure for excel sheet, used both on import and export''' 

    def __init__(self,  ws):
        self.ws = ws
        self.cols = {}
        c = 0
        for t in ["Aktivitet", "Typ", "Datum", "Vecka", "Dag",
                  "Uppgift", "Start", "Slut",
                  "Ersättning", "Mat", "Hyrkart",
                  "Publ. datum", "Koordinator",
                  "Guldkortsnr", "Namn", "Mobilnummer", "Kommentar"]:
            self.cols[t] = c        
            c += 1

    def createHeader(self, row = 1):
        for t, c in self.cols.items():
            ws.cell(row, c).value = t
            ws.column_dimensions[chr(64+c)].best_fit = True

class ExcelRow:
    def __init__(self, sheet, row):
        self.sheet = sheet
        self._row = row

    def cell(self, header):
        return self._row[self.sheet.cols[header]]

    @property
    def event(self):
        return self.cell("Aktivitet")

    @property
    def event_type(self):
        return self.cell("Typ")

    @property
    def date(self):
        return self.cell("Datum")

    @property
    def activity(self):
        return self.cell("Uppgift")

    @property
    def start_time(self):
        return self.cell("Start")

    @property
    def end_time(self):
        return self.cell("Slut")

    @property
    def reimbursement(self):
        return self.cell("Ersättning")

    @property
    def food(self):
        return self.cell("Mat")

    @property
    def rental_kart(self):
        return self.cell("Hyrkart")

    @property
    def bookable_date(self):
        return self.cell("Publ. datum")

    @property
    def coordinator(self):
        return self.cell("Koordinator")



def importDataFromExcel(file, year=datetime.date.today().year):
    wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
    sheet = wb._sheets[0]

    locale.setlocale(locale.LC_ALL, 'sv_SE')

    events = {}
    eventTypes = {}
    activityTypes = {}
    n = 0
    et_name = None

    try:
        es = ExcelSheet(sheet)

        for cells in sheet.rows:
            row = ExcelRow(es, cells)
            n += 1
            if n <= 2:
                continue

            if row.event.value is None and et_name is None:
                continue

            et_name = row.event_type.value or et_name
            et = eventTypes.get(et_name)
            if et is None:
                #logger.info(f"Event type {et_name}")
                try:
                    et = EventType.objects.get(name=et_name)
                except EventType.DoesNotExist:
                    et = EventType(name=et_name)
                    et.save()
                eventTypes[et_name] = et

            if row.date.value is None:
                continue

            date = row.date.value
            if year is not None:
                (_, week, weekday) = date.isocalendar()            
                date = datetime.date.fromisocalendar(year, week, weekday)
            else:
                (_, week, weekday) = date.isocalendar()            

            if any(et_name.startswith(t) for t in ['Träning', 'Arbetsdag', 'Gokartskola']):
                event_name = f"{et_name} {calendar.day_name[weekday - 1]} vecka {week}"            
            else:    
                event_name = row.event.value

            event = events.get((event_name, date))

            if event is None:
                #logger.info(f"Event {event_name} {date}")
                try:
                    event = Event.objects.get(name=event_name,
                                              start_date=date, type=et)
                except Event.DoesNotExist:
                    event = Event(name=event_name, start_date=date,
                                  end_date=date, type=et)
                    event.save()

                    coord = row.coordinator.value

                    if coord is not None:
                        if ' ' in coord:
                            first, last = coord.split(' ', maxsplit=1)
                        else:
                            first = coord

                        try:
                            coordinator = Member.objects.get(
                                user__first_name=first, user__last_name=last)
                            event.coordinators.add(coordinator)
                        except Member.DoesNotExist:
                            print(
                                f"Failed to find member {coord} to use as coordinator for {event_name}")

                events[event_name] = event

            at_name = row.activity.value
            if at_name is None or at_name == '':
                print(f"Skipping row {n} as 'uppgift' is empty")
                continue

            at = activityTypes.get(at_name)
            if at is None:
                #logger.info(f"Activity Type: {at_name}")
                try:
                    at = ActivityType.objects.get(name=at_name)
                except ActivityType.DoesNotExist:
                    at = ActivityType(name=at_name)
                    at.save()
                activityTypes[at_name] = at

            ebd = row.bookable_date.value

            activity = Activity(
                name=f"{at_name} {calendar.day_name[date.weekday()]}", event=event, type=at,
                earliest_bookable_date=ebd)

            activity.start_time = row.start_time.value
            activity.end_time = row.end_time.value

            activity.full_clean()
            activity.save()

        print(f'''Database row count:
            {EventType.objects.all().count()} event types
            {ActivityType.objects.all().count()} activity types
            {Event.objects.all().count()} events
            {Activity.objects.all().count()} activities'''
              .replace('    ', ' '))

    except:
        print(f"Error on row {n}")
        raise   

def exportScheduleToExcel(filename, year):
    wb = openpyxl.Workbook(iso_dates=True)
    ws = wb._sheets[0]

    driver_type = ActivityType.objects.get(name__startswith="Förare")

    es = ExcelSheet(ws)
    es.createHeader()

    cols = es.cols

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['F'].width = 25
    for c in ['B','C','D','E','L']:
        ws.column_dimensions[c].width = 10
    for c in ['G','H']:
        ws.column_dimensions[c].width = 7
    for c in ['I','J','K']:
        ws.column_dimensions[c].width = 4

    header_font = openpyxl.styles.Font(bold=True)
    computed_font = openpyxl.styles.Font(italic=True)

    for cell in ws["1:1"]:
        cell.font = header_font

    r = 3
    for event in Event.objects.filter(start_date__year=year):
        print(f'{event.start_date}: {event.name}')

        ws.cell(r, 1).value = event.name
        ws.cell(r, cols['Typ']).value = event.type.name if event.type else ''

        datecell = ws.cell(r, cols['Datum'])
        datecell.value = event.start_date
        datecell.number_format = 'YYYY-MM-DD'

        weekcell = ws.cell(r, cols['Vecka'])
        weekcell.value = f'=ISOWEEKNUM({datecell.coordinate})' 
        weekcell.data_type = 'f'

        daycell = ws.cell(r, cols['Dag'])
        daycell.value = f'={datecell.coordinate}' 
        daycell.number_format = 'ddd'

        for c in [weekcell, daycell]:
            c.font = computed_font

        ws.cell(r, cols['Koordinator']).value = ", ".join(
            c.get_fullname() for c in event.coordinators.all())

        for a in event.activities.exclude(type__in=[driver_type]).order_by('type__id'):
            col = [a.start_time, a.end_time]
            for i, t in enumerate(col):
                ws.cell(r, cols['Start'] + i).value = t
                ws.cell(r, cols['Start'] + i).number_format = 'hh:mm'

            ws.cell(r, cols['Publ. datum']).value = a.earliest_bookable_date

            if a.type:
                ws.cell(r, cols['Uppgift']).value = a.type.name
                for i, t in enumerate([ a.type.fee_reimbursed,
                                   a.type.food_included, a.type.rental_kart]):
                    ws.cell(r, cols['Ersättning'] + i).value = 'Ja' if t else 'Nej' 

            r += 1


    with open(filename, 'wb') as file:
        wb.save(file)

    print(f"Exported {r-2} activities to: {filename}")
